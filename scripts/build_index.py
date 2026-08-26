#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""사양서 엑셀 폴더를 훑어 색인(JSON)을 만든다.

사내에서 돌리는 도구다. 사이트는 이 색인만 읽고 엑셀은 열지 않는다.
폴더가 원본이고 색인은 파생물이다 — 둘이 어긋나면 폴더가 옳다.

    python3 scripts/build_index.py "//서버/공유/사양서 모음" -o js/index-data.js

기획자가 보낸 10건을 실제로 파싱해 보고 쓴 규칙이다. 그 10건이 이미
서로 다른 양식이었다 (CLAUDE.md §3). 요약하면:

  · 시트 이름이 7종이라 시트를 이름으로 고를 수 없다
  · 제목 줄이 밀려 품번이 B3 인 파일과 B4 인 파일이 섞여 있다
  · 같은 항목을 다른 이름으로 부른다 (메이커 / 제조 Maker)
  · 사양서가 아닌 문서(구매 목록)가 같은 폴더에 있다

그래서 **위치를 믿지 않는다.** 라벨을 찾고 그 오른쪽 칸을 읽는다.
"""
import argparse
import json
import os
import re
import sys
import unicodedata

# ─────────────────────────────────────────── 항목 정의
#
# 라벨은 파일마다 다르게 적혀 있다. 왼쪽이 우리가 쓸 이름, 오른쪽이
# 실제로 발견된 표기들. 새 표기를 만나면 여기에 더한다 —
# 파서 본문을 고치는 것이 아니라 이 표를 늘리는 것이 정상적인 유지보수다.
FIELDS = [
    ('partNo',     '품번',        ['품번', '품 번', 'PARTNO', 'PART NO']),
    ('name',       '품명',        ['품명', '품 명', 'PARTNAME', 'PART NAME']),
    ('model',      '모델 및 규격', ['모델및규격', '규격및모델', '규격', '모델']),
    ('maker',      '제조 Maker',  ['제조maker', '메이커', 'maker', '제조사']),
    ('use',        '용도',        ['용도']),
    ('material',   '재질',        ['재질']),
    ('unit',       '구매단위',    ['구매단위']),
    ('qtyPerUnit', '단위당 수량', ['단위당수량']),
    ('origin',     '원산지',      ['원산지']),
    ('detail',     '상세규격',    ['상세규격', '부품설명']),
    ('remark',     '비고',        ['remark', '별도요청사항', '비고']),
    ('bg',         'B G',        ['bg', 'b g']),
    ('contact',    '주문자 연락처', ['주문자연락처', '주문자', '연락처']),
]
LABEL_TO_KEY = {}
for key, _title, aliases in FIELDS:
    for a in aliases:
        LABEL_TO_KEY.setdefault(norm_seed := re.sub(r'\s+', '', a).lower(), key)

FIELD_TITLE = {k: t for k, t, _ in FIELDS}

# 품번은 6자리-5자리다. 12번 프로젝트(6-6)와 다르므로 그대로 옮겨 쓰면 안 된다.
PART_NO = re.compile(r'\b(\d{6}-\d{5})\b')

# 문서 종류 — 제목 칸으로 가른다
KINDS = [
    ('spec',     ['부품사양서', 'specsheet', '부품사양']),
    ('desc',     ['부품설명']),
    ('purchase', ['구매목록', '구매리스트']),
]


def norm(s):
    """공백·전각·대소문자를 지우고 비교용으로 만든다.

    엑셀에서 라벨은 '품   번' 처럼 글자 사이에 공백을 넣어 폭을 맞춘 것이 많다.
    전각 공백(　)도 실제로 섞여 있었다.
    """
    if s is None:
        return ''
    s = unicodedata.normalize('NFKC', str(s))
    return re.sub(r'\s+', '', s).lower()


# ─────────────────────────────────────────── 엑셀 읽기
def read_cells(path):
    """{'B4': '품번', ...} 로 첫 시트를 읽는다. .xls 와 .xlsx 를 모두 받는다."""
    if path.lower().endswith(('.xlsx', '.xlsm')):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        cells = {}
        for row in ws.iter_rows(min_row=1, max_row=40, max_col=14):
            for c in row:
                if c.value not in (None, ''):
                    cells[c.coordinate] = str(c.value).strip()
        return wb.sheetnames[0], cells
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    cells = {}
    for r in range(min(ws.nrows, 40)):
        for c in range(min(ws.ncols, 14)):
            v = ws.cell_value(r, c)
            if v not in (None, ''):
                cells['%s%d' % (chr(65 + c), r + 1)] = str(v).strip()
    return ws.name, cells


def split_ref(ref):
    m = re.match(r'([A-Z]+)(\d+)', ref)
    return m.group(1), int(m.group(2))


def col_next(col, n=1):
    """B → C. 두 글자 열(AA)은 이 양식에 나오지 않지만 넘어와도 죽지 않게 둔다."""
    if len(col) != 1:
        return col
    return chr(ord(col) + n)


def find_field(cells, key):
    """라벨을 찾아 그 **오른쪽** 칸을 읽는다.

    위치(C4 등)를 쓰지 않는 이유가 이 함수의 존재 이유다.
    같은 폴더의 파일들이 이미 한 칸씩 밀려 있었다.

    돌려주는 값은 세 가지로 갈린다. 이 구분이 규칙 5다.
      (None,  'missing') 항목(라벨) 자체가 이 사양서에 없다
      ('',    'blank')   라벨은 있는데 칸이 비어 있다
      ('값',  'filled')

    '없음'과 '빈칸'을 뭉뚱그리면 화면이 둘 다 '-' 로 보이고,
    사양서에 원래 없는 항목인지 담당자가 안 적은 것인지 알 수 없게 된다.

    **같은 라벨이 여러 번 나오면 전부 담는다.** 실제 파일에 '용 도' 가
    머리줄에 한 번(짧게 '마스킹용'), 본문에 한 번(길게 '- 고무자석입니다…')
    나오는 양식이 있었다. 앞엣것만 잡으면 검색이 걸려야 할 긴 문장을
    통째로 버린다. 어느 쪽이 진짜인지 파서가 고를 일이 아니다.
    """
    found = False
    values = []
    for ref, text in sorted(cells.items(), key=lambda kv: (split_ref(kv[0])[1], split_ref(kv[0])[0])):
        if LABEL_TO_KEY.get(norm(text)) != key:
            continue
        found = True
        col, row = split_ref(ref)
        # 오른쪽으로 두 칸까지 본다 — 라벨 옆에 병합된 빈 칸을 두는 양식이 있었다
        # (품번 라벨이 B3, 값이 D3 인 파일이 실제로 있다)
        for step in (1, 2):
            v = cells.get('%s%d' % (col_next(col, step), row))
            if v in (None, ''):
                continue
            # 오른쪽 칸이 또 다른 라벨이면 이 칸은 비어 있는 것이다
            # (구매목록의 머리줄: 품번 | 품명 | 규격 …)
            if LABEL_TO_KEY.get(norm(v)):
                break
            v = v.strip()
            if v and v not in values:
                values.append(v)
            break
    if not found:
        return None, 'missing'
    if not values:
        return '', 'blank'
    return '\n'.join(values), 'filled'


def classify(cells):
    """문서 종류. 사양서가 아닌 것을 사양서로 내주지 않기 위해 있다."""
    for ref, text in sorted(cells.items(), key=lambda kv: (split_ref(kv[0])[1], kv[0])):
        n = norm(text)
        for kind, keys in KINDS:
            if any(k in n for k in keys):
                return kind
    return 'unknown'


def resolve_part_no(file_no, body_no):
    """파일명의 품번과 내용의 품번을 맞춰 본다.

    둘이 다를 때 **한쪽을 조용히 고르지 않는다.** 10건 중 2건이 어긋났다.
    - 내용이 비어 있으면 파일명을 쓰되 출처를 남긴다
    - 둘 다 있는데 다르면 파일명을 쓰고 conflict 로 표시한다
      (폴더에서 파일을 찾는 것은 파일명으로 하기 때문이다)
    """
    if body_no and file_no and body_no != file_no:
        return file_no, 'conflict'
    if body_no:
        return body_no, 'body'
    if file_no:
        return file_no, 'filename'
    return '', 'none'


def parse(path):
    sheet, cells = read_cells(path)
    base = os.path.basename(path)
    m = PART_NO.search(base)
    file_no = m.group(1) if m else ''

    body_raw, body_state = find_field(cells, 'partNo')
    body_no = ''
    if body_state == 'filled':
        m2 = PART_NO.search(body_raw)
        body_no = m2.group(1) if m2 else ''

    part_no, source = resolve_part_no(file_no, body_no)
    rec = {
        'partNo': part_no,
        'partNoSource': source,
        'kind': classify(cells),
        'file': base,
        'sheet': sheet,
        'fields': {},
    }
    for key, _title, _ in FIELDS:
        if key == 'partNo':
            continue
        value, state = find_field(cells, key)
        rec['fields'][key] = {'value': value or '', 'state': state}
    return rec


# ─────────────────────────────────────────── 최신화 비교
#
# 색인을 다시 만들 때 무엇이 달라졌는지 말한다.
# "뭔가 바뀐 것 같다"가 아니라 **어느 칸이 무엇에서 무엇으로** 바뀌었는지
# 말할 수 있어야, 그 변경이 맞는지 사람이 판단할 수 있다.
#
# ⚠ js/logic.js 의 compareIndex() 와 **같은 규칙**이다. 한쪽만 고치지 말 것.
COMPARE_FIELDS = ['name', 'model', 'maker', 'use', 'material',
                  'unit', 'qtyPerUnit', 'origin', 'detail', 'remark', 'bg']


def _cut(v, n=40):
    v = re.sub(r'\s*\n+\s*', ' ', str(v or '')).strip()
    return v[:n] + '…' if len(v) > n else v


def diff_record(before, after):
    out = []
    if before.get('partNo') != after.get('partNo'):
        out.append(('품번', before.get('partNo', ''), after.get('partNo', '')))
    if before.get('kind') != after.get('kind'):
        out.append(('문서 종류', before.get('kind', ''), after.get('kind', '')))
    for k in COMPARE_FIELDS:
        a = (before.get('fields') or {}).get(k) or {'value': '', 'state': 'missing'}
        b = (after.get('fields') or {}).get(k) or {'value': '', 'state': 'missing'}
        if a.get('value') == b.get('value') and a.get('state') == b.get('state'):
            continue
        out.append((FIELD_TITLE.get(k, k),
                    _cut(a.get('value')) or '(%s)' % a.get('state'),
                    _cut(b.get('value')) or '(%s)' % b.get('state')))
    return out


def compare_index(before, after):
    a = {r['file']: r for r in before if r.get('file')}
    b = {r['file']: r for r in after if r.get('file')}
    out = {'added': [], 'changed': [], 'same': [], 'removed': []}
    for rec in after:
        old = a.get(rec.get('file'))
        if old is None:
            out['added'].append(rec)
            continue
        d = diff_record(old, rec)
        if d:
            out['changed'].append((rec, d))
        else:
            out['same'].append(rec)
    for f, rec in a.items():
        if f not in b:
            out['removed'].append(rec)
    return out


def print_compare(cmp):
    print('')
    print('최신화 비교 — 신규 %d · 변경 %d · 동일 %d · 사라짐 %d' % (
        len(cmp['added']), len(cmp['changed']), len(cmp['same']), len(cmp['removed'])))
    for rec in cmp['added']:
        print('  [신규]   %s  %s' % (rec.get('partNo', '?'), rec.get('file', '')))
    for rec, diffs in cmp['changed']:
        print('  [변경]   %s  %s' % (rec.get('partNo', '?'), rec.get('file', '')))
        for title, a, b in diffs:
            print('           · %s 변경: %s → %s' % (title, a or '(없음)', b or '(없음)'))
    for rec in cmp['removed']:
        # 지우지 않고 알린다. 파일을 옮겼을 수도 있고 실수로 지웠을 수도 있다.
        print('  [사라짐] %s  %s' % (rec.get('partNo', '?'), rec.get('file', '')))
    if not (cmp['added'] or cmp['changed'] or cmp['removed']):
        print('  달라진 것이 없습니다.')


# ─────────────────────────────────────────── 실행
def walk(folder):
    out = []
    for root, _dirs, files in os.walk(folder):
        for f in sorted(files):
            if f.startswith('~$') or not f.lower().endswith(('.xls', '.xlsx', '.xlsm')):
                continue
            p = os.path.join(root, f)
            try:
                out.append(parse(p))
            except Exception as e:            # 한 건이 깨져도 나머지는 만든다
                out.append({'partNo': '', 'partNoSource': 'none', 'kind': 'error',
                            'file': f, 'sheet': '', 'fields': {},
                            'error': '%s: %s' % (type(e).__name__, e)})
    return out


def main(argv=None):
    ap = argparse.ArgumentParser(description='사양서 엑셀 폴더 → 색인')
    ap.add_argument('folder', help='사양서가 들어 있는 폴더')
    ap.add_argument('-o', '--out', default='js/index-data.js')
    ap.add_argument('--json', action='store_true', help='JS 대신 순수 JSON 으로')
    ap.add_argument('--compare', metavar='이전색인.json',
                    help='기존 색인과 견줘 신규/변경/동일을 알려 준다')
    a = ap.parse_args(argv)

    recs = walk(a.folder)
    body = json.dumps(recs, ensure_ascii=False, indent=1)
    if a.json:
        text = body
    else:
        text = ('/* scripts/build_index.py 가 만든 파일 — 손으로 고치지 마세요. */\n'
                'window.SPEC_INDEX = %s;\n' % body)
    with open(a.out, 'w', encoding='utf-8') as f:
        f.write(text)

    kinds = {}
    for r in recs:
        kinds[r['kind']] = kinds.get(r['kind'], 0) + 1
    conflict = [r for r in recs if r['partNoSource'] == 'conflict']
    print('%d 건 → %s' % (len(recs), a.out))
    print('  종류: ' + ', '.join('%s %d' % kv for kv in sorted(kinds.items())))
    if conflict:
        print('  ⚠ 파일명과 내용의 품번이 다른 파일 %d 건:' % len(conflict))
        for r in conflict:
            print('     %s' % r['file'])
    if a.compare:
        try:
            with open(a.compare, encoding='utf-8') as f:
                before = json.load(f)
        except Exception as e:
            print('  ⚠ 이전 색인을 읽지 못했습니다: %s' % e)
        else:
            print_compare(compare_index(before, recs))

    not_spec = [r for r in recs if r['kind'] not in ('spec', 'desc')]
    if not_spec:
        print('  ⚠ 사양서가 아닌 파일 %d 건 (검색 결과로 내주지 않습니다):' % len(not_spec))
        for r in not_spec:
            print('     %s [%s]' % (r['file'], r['kind']))
    return 0


if __name__ == '__main__':
    sys.exit(main())
